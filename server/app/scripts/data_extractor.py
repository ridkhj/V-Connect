import re
import openpyxl
import os
from app.components.cdpr import Cdpr
from app.components.update import Update
from app.components.letter import Letter
from app.components.reciprocal_letter import ReciprocalLetter
from app.components.nsl_letter import NslLetter
from app.components.thankyou_letter import ThankyouLetter
from pathlib import Path 



class SheetDataExtractor:

    def __init__(self):
        self._path = Path(__file__).parent.parent / 'assets'
        self._updates: list[Update] = []
        self._cdprs: list[Cdpr] = []
        self._reciprocal_letters: list[ReciprocalLetter] = []
        self._nsl_letters: list[NslLetter] = []
        self._thankyou_letters: list[ThankyouLetter] = []
        
    @property
    def path(self):
        return self._path

    @path.setter
    def path(self, value):
        self._path = value

    @property
    def atualizacoes(self):
        return self._atualizacoes

    @atualizacoes.setter
    def atualizacoes(self, value: list[Update]):
        self._atualizacoes = value

    @property
    def cdprs(self):
        return self._cdprs

    @cdprs.setter
    def cdprs(self, value: list[Cdpr]):
        self._cdprs = value
    
    @property
    def letters(self):
        return self._letters
    
    @letters.setter
    def letters(self, value: list[Letter]):
        self._letters = value

    def unity_find_sheet(self, name): #recebe nome por exemplo 'atualizacoes.xlsx' ou 'cdpr.xlsx'
        loadPath = self.path / 'sheets' / name

        if not loadPath.exists():
            raise FileNotFoundError(f"O arquivo {loadPath} não foi encontrado")

        return loadPath
    
    def unity_process_sheet(self, fileName):
        
        filePath = self.unity_find_sheet(fileName)

        if fileName.startswith('atualizacoes'):
            return self.extract_updates_data(filePath)
        elif fileName.startswith('cdpr'): 
            return self.extract_cdpr_data(filePath) 
        elif fileName.startswith('reciprocas'):
            return self.extract_reciprocal_letter_data(filePath)
        elif fileName.startswith('nsl'):
            return self.extract_nsl_data(filePath)
        elif fileName.startswith('agradecimento'):
            return self.extract_thankyou_letter_data(filePath)


    def select_and_process_sheet_by_type(self): 
        loadPath = self.path / 'sheets'

        if loadPath.exists():

            files = os.listdir(loadPath)
            files_xlsx = [file for file in files if file.endswith('.xlsx')]

            for file in files_xlsx:
                self.unity_process_sheet(file)
        else:
            raise FileNotFoundError(f"O diretório {loadPath} não foi encontrado.")
                    
    def extract_updates_data(self, path):
        file = openpyxl.load_workbook(path)
        sheet = file.active
        atualizacoes = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        access = sheet.cell

        while (access(row, column).value is not None and access(row, column+1).value is not None):

            codigo = access(row, column).value
            nome  = access(row, column+1).value
            status = access(row, column+3).value

            
            atualizacaoAux = Update(codigo, nome, status)

            if status != "Enviado" and status != "Devolvido à Igreja Parceira":
        
                atualizacoes.append(atualizacaoAux)
                
            row += 1

        self.updates = atualizacoes
        return self.updates

    def extract_cdpr_data(self, path):
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        cdprs = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        access = planilha.cell

        while (access(row, column).value is not None and access(row, column+1).value is not None):

            nome = access(row, column).value
            codigo = access(row, column+1).value
            age = access(row, column+2).value

            cdprs.append(Cdpr(codigo, nome, age))
    
            row += 1

        self.cdprs = cdprs
        return self.cdprs
    
    def extract_reciprocal_letter_data(self, path):
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        letters = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        access = planilha.cell


        while (access(row, column).value is not None and access(row, column+1).value is not None):

            code = access(row, column).value
            letterCode = access(row, column+1).value
            name = access(row, column+2).value
            ageBracket = access(row, column+3).value
            type = access(row, column+5).value
            questions = access(row, column+6).value
            status = access(row, column+12).value   

            
            letters.append(ReciprocalLetter(code = code, letterCode= letterCode, name=name,type= type,questions= questions,status= status,ageBracket= ageBracket))


            row += 1

        self._reciprocal_letters.extend(letters)
        return self._reciprocal_letters
    
    def extract_nsl_data(self, path):
        
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        letters = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        access = planilha.cell
      

        while (access(row, column).value is not None and access(row, column+1).value is not None):

            code = access(row, column).value
            letterCode = access(row, column+1).value
            name = access(row, column+2).value
            type = access(row, column+4).value
            status = access(row,12).value   

            letters.append(NslLetter(code=code,letterCode= letterCode,name= name,type= type,status= status))

          

            row += 1

        self._nsl_letters.extend(letters)
        return self._nsl_letters

    def extract_thankyou_letter_data(self, path):
         
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        letters = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        access = planilha.cell
       
        while (access(row, column).value is not None and access(row, column+1).value is not None):

            code = access(row, column).value
            letterCode = access(row, column+1).value
            name = access(row, column+2).value
            type = access(row, column+4).value
            questions = access(row, column+10).value
            status = access(row,column+12).value   

            letters.append(ThankyouLetter(code=code,letterCode= letterCode,name= name,type= type,status= status, questions=questions))
            
            row += 1

        self._thankyou_letters.extend(letters)
        return self._thankyou_letters
        