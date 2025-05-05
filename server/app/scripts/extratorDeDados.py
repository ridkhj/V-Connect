import openpyxl
import os
from components.cdpr import Cdpr
from components.atualizacao import Atualizacao
from components.carta import CartaNsl, CartaReciproca, Carta, CartaAgradecimento
from pathlib import Path 



class ExtratorDeDados:

    def __init__(self):
        self._path = Path(__file__).parent.parent / 'assets'
        self._atualizacoes: list[Atualizacao] = []
        self._cdprs: list[Cdpr] = []
        self._cartas: list[Carta] = []
        
    @property
    def path(self):
        return self._path

    @path.setter
    def path(self, value):
        self._path = value

    @property
    def atualizacoes(self):
        return self._atualizacoes

    @atualizacoes.setter
    def atualizacoes(self, value: list[Atualizacao]):
        self._atualizacoes = value

    @property
    def cdprs(self):
        return self._cdprs

    @cdprs.setter
    def cdprs(self, value: list[Cdpr]):
        self._cdprs = value
    
    @property
    def cartas(self):
        return self._cartas
    
    @cartas.setter
    def cartas(self, value: list[Carta]):
        self._cartas = value

    def listarArquivos(self):
        loadPath = self.path / 'sheets'
        if loadPath.exists():
            arquivos = os.listdir(loadPath)
            arquivos_xlsx = [arquivo for arquivo in arquivos if arquivo.endswith('.xlsx')]
            for arquivo in arquivos_xlsx:
                if arquivo.startswith('atualizacoes'):
                    self.extrairDadosAtualizacao(os.path.join(loadPath, arquivo))
                elif arquivo.startswith('cdpr'): 
                    self.extrairDadosCDPR(os.path.join(loadPath, arquivo)) 
                elif arquivo.startswith('reciprocas'):
                    self.extrairDadosReciprocas(os.path.join(loadPath, arquivo))
                elif arquivo.startswith('nsl'):
                    self.extrairDadosNsl(os.path.join(loadPath, arquivo))
                elif arquivo.startswith('agradecimento'):
                    self.extrairDadosAgradecimento(os.path.join(loadPath, arquivo))
        else:
            raise FileNotFoundError(f"O diretório {loadPath} não foi encontrado.")
                
    def extrairDadosAtualizacao(self, path):
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        atualizacoes = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        acessar = planilha.cell

        while (acessar(row, column).value is not None and acessar(row, column+1).value is not None):

            codigo = acessar(row, column).value
            nome = acessar(row, column+1).value
            status = acessar(row, column+3).value

            
            atualizacaoAux = Atualizacao(codigo, nome, status)

            if status != "Enviado" and status != "Devolvido à Igreja Parceira":
        
                atualizacoes.append(atualizacaoAux)
                
            row += 1

        self.atualizacoes = atualizacoes

    def extrairDadosCDPR(self, path):
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        cdprs = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        acessar = planilha.cell

        while (acessar(row, column).value is not None and acessar(row, column+1).value is not None):

            codigo = acessar(row, column).value
            nome = acessar(row, column+1).value
            age = acessar(row, column+3).value

            cdprs.append(Cdpr(codigo, nome, age))
    
            row += 1

        self.cdprs = cdprs
    
    def extrairDadosReciprocas(self, path):
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        cartas = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        acessar = planilha.cell


        while (acessar(row, column).value is not None and acessar(row, column+1).value is not None):

            code = acessar(row, column).value
            letterCode = acessar(row, column+1).value
            name = acessar(row, column+2).value
            ageBracket = acessar(row, column+3).value
            type = acessar(row, column+5).value
            questions = acessar(row, column+6).value
            status = acessar(row, column+12).value   

            
            cartas.append(CartaReciproca(code = code, letterCode= letterCode, name=name,type= type,questions= questions,status= status,ageBracket= ageBracket))


            row += 1

        self._cartas.extend(cartas)
    
    def extrairDadosNsl(self, path):
        
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        cartas = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        acessar = planilha.cell
      

        while (acessar(row, column).value is not None and acessar(row, column+1).value is not None):

            code = acessar(row, column).value
            letterCode = acessar(row, column+1).value
            name = acessar(row, column+2).value
            type = acessar(row, column+4).value
            status = acessar(row,12).value   

            cartas.append(CartaNsl(code=code,letterCode= letterCode,name= name,type= type,status= status))

          

            row += 1

        self._cartas.extend(cartas)

    def extrairDadosAgradecimento(self, path):
         
        arquivo = openpyxl.load_workbook(path)
        planilha = arquivo.active
        cartas = []

        row = 2        # coordenadas da primeira celula a ser lida 
        column = 2
        acessar = planilha.cell
       
        while (acessar(row, column).value is not None and acessar(row, column+1).value is not None):

            code = acessar(row, column).value
            letterCode = acessar(row, column+1).value
            name = acessar(row, column+2).value
            type = acessar(row, column+4).value
            questions = acessar(row, column+9).value
            status = acessar(row,column+10).value   

            cartas.append(CartaAgradecimento(code=code,letterCode= letterCode,name= name,type= type,status= status, questions=questions))
            
            row += 1

        self._cartas.extend(cartas)